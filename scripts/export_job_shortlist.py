"""Export the prioritized job shortlist and adaptation status to XLSX."""

from __future__ import annotations

import json
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
JOBS_DIR = ROOT / "data" / "jobs"
OUTPUT = ROOT / "vagas_priorizadas.xlsx"

JOB_IDS = (
    "olist-especialista-gtm-erp",
    "xtrategus-ai-productivity-adoption-consultant",
    "nuvemshop-sales-partner-enablement-specialist",
    "bairesdev-revenue-operations-specialist",
    "grupo-ric-especialista-sales-enablement",
    "doctoralia-noa-gtm-revenue-operations-specialist",
    "playlist-revenue-enablement-specialist",
    "modaxo-ai-transformation-manager",
    "rd-station-revops-senior-abm-ia",
    "infobip-senior-partnership-enablement-specialist",
    "next-level-growth-marketing-operations-project-manager",
    "revblack-revops-consultant",
    "wellhub-gtm-program-manager-revenue-operations",
    "ebanx-revenue-strategy-mid-analyst",
    "techne-especialista-automacao-ia",
)


def read_job(job_id: str) -> dict[str, Any]:
    with (JOBS_DIR / f"{job_id}.json").open(encoding="utf-8") as file:
        return json.load(file)


def join_text(values: list[Any] | None) -> str:
    return "\n".join(str(value) for value in values or [] if value)


def recommendation(fit_score: int) -> str:
    if fit_score >= 90:
        return "Aplicar imediatamente"
    if fit_score >= 82:
        return "Boa aderência"
    return "Stretch / validar knockouts"


def unique_text(*groups: list[Any] | None) -> str:
    values: list[str] = []
    seen: set[str] = set()
    for group in groups:
        for value in group or []:
            text = str(value).strip()
            if text and text.casefold() not in seen:
                seen.add(text.casefold())
                values.append(text)
    return "\n".join(values)


def skill_text(job: dict[str, Any]) -> str:
    skills: list[str] = []
    for category in job["resume"]["pt"].get("skills", []):
        skills.extend(category.get("skills", []))
    return ", ".join(dict.fromkeys(skills))


def build_workbook(jobs: list[dict[str, Any]]) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Vagas priorizadas"

    headers = [
        "Ranking",
        "Empresa",
        "Vaga",
        "Fit (%)",
        "Recomendação",
        "Local",
        "Modelo",
        "Contrato",
        "Idioma do anúncio",
        "Remuneração",
        "Prazo",
        "Pontos fortes",
        "Lacunas e riscos ATS",
        "Palavras-chave do currículo",
        "Link da vaga",
        "Artefato versionado",
        "Currículo PT (PDF)",
        "Currículo EN (PDF)",
        "Carta PT (PDF)",
        "Carta EN (PDF)",
        "Candidatura",
    ]
    sheet.append(headers)

    jobs.sort(key=lambda job: job["metadata"]["fit_score"], reverse=True)
    for rank, job in enumerate(jobs, start=1):
        metadata = job["metadata"]
        triage = job["triage"]
        compensation = (
            metadata.get("compensation")
            or metadata.get("salary_expectation")
            or "Não informada"
        )
        deadline = (
            metadata.get("application_deadline")
            or metadata.get("deadline")
            or triage.get("deadline")
            or "Não informado"
        )
        sheet.append(
            [
                rank,
                metadata["company_name"],
                metadata["role_title"],
                metadata["fit_score"],
                recommendation(metadata["fit_score"]),
                metadata.get("location", "Não informado"),
                metadata.get("work_model", "Não informado"),
                metadata.get("employment_type", "Não informado"),
                metadata.get("document_language", "pt").upper(),
                compensation,
                deadline,
                join_text(metadata.get("good_points")),
                unique_text(
                    triage.get("blockers"),
                    triage.get("gaps"),
                    triage.get("risks"),
                    metadata.get("improvement_points"),
                ),
                skill_text(job),
                metadata["url"],
                f"data/jobs/{job['id']}.json",
                f"exports/curriculos_vagas/{job['id']}/curriculo_pt.pdf",
                f"exports/curriculos_vagas/{job['id']}/curriculo_en.pdf",
                f"exports/curriculos_vagas/{job['id']}/carta_pt.pdf",
                f"exports/curriculos_vagas/{job['id']}/carta_en.pdf",
                "Não enviada",
            ]
        )

    header_fill = PatternFill("solid", fgColor="17324D")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        link_cell = row[14]
        link_cell.hyperlink = link_cell.value
        link_cell.style = "Hyperlink"
        for file_cell in row[16:20]:
            file_cell.hyperlink = file_cell.value
            file_cell.style = "Hyperlink"

    widths = {
        "A": 9,
        "B": 22,
        "C": 38,
        "D": 10,
        "E": 27,
        "F": 22,
        "G": 22,
        "H": 22,
        "I": 18,
        "J": 35,
        "K": 18,
        "L": 55,
        "M": 65,
        "N": 55,
        "O": 55,
        "P": 52,
        "Q": 20,
        "R": 20,
        "S": 20,
        "T": 20,
        "U": 18,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.row_dimensions[1].height = 32
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    table = Table(displayName="VagasPriorizadas", ref=sheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.conditional_formatting.add(
        f"D2:D{sheet.max_row}",
        ColorScaleRule(
            start_type="num",
            start_value=60,
            start_color="F8696B",
            mid_type="num",
            mid_value=82,
            mid_color="FFEB84",
            end_type="num",
            end_value=100,
            end_color="63BE7B",
        ),
    )

    guide = workbook.create_sheet("Guia")
    guide_rows = [
        ("Planilha", "Vagas priorizadas e currículos adaptados"),
        ("Atualizada em", date.today().isoformat()),
        ("Total de vagas", len(jobs)),
        ("Status", "Todos os currículos PT/EN foram criados e validados"),
        ("Candidaturas", "Nenhuma candidatura foi marcada como enviada"),
        ("Aplicar imediatamente", "Fit de 90% ou mais"),
        ("Boa aderência", "Fit entre 82% e 89%"),
        ("Stretch", "Fit abaixo de 82%; revisar knockouts antes de aplicar"),
        ("Fonte", "master_resume.json; sem inclusão de experiência não comprovada"),
    ]
    for row in guide_rows:
        guide.append(row)
    guide.column_dimensions["A"].width = 25
    guide.column_dimensions["B"].width = 90
    for cell in guide["A"]:
        cell.font = Font(bold=True, color="17324D")
    for row in guide.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    return workbook


def main() -> None:
    jobs = [read_job(job_id) for job_id in JOB_IDS]
    build_workbook(jobs).save(OUTPUT)
    print(f"Exported {len(jobs)} jobs to {OUTPUT}")


if __name__ == "__main__":
    main()
